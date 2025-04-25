"""Handles the generation of optimization summary Excel files."""

import os
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any

# Added imports
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill  # Added for cell styling

# Define constants for requested stats to avoid repetition
REQUESTED_STATS = [
    "Equity Final [$]",
    "Commissions [$]",
    "Return [%]",
    "Return (Ann.) [%]",
    "Max. Drawdown [%]",
    "Max. Drawdown Duration",
    "# Trades",
    "Win Rate [%]",
    "Sharpe Ratio",
    "Sortino Ratio",
    "Calmar Ratio",
    "SQN",
    "Kelly Criterion",
    "Profit Factor",
]


def _process_results_to_dataframe(run_results: List[Dict[str, Any]]) -> pd.DataFrame:
    """
    Processes a list of run results into a structured Pandas DataFrame,
    dynamically including optimized parameters.

    Args:
        run_results: List of result dictionaries.

    Returns:
        A Pandas DataFrame with the processed results.
    """
    summary_data_for_excel = []
    all_best_param_keys = set()

    # First pass: Collect all unique keys from best_params across all results
    for run_result in run_results:
        best_params = run_result.get("best_params", {})
        all_best_param_keys.update(best_params.keys())

    # Sort parameter keys for consistent column order
    sorted_best_param_keys = sorted(list(all_best_param_keys))

    # Second pass: Build the data for the DataFrame
    for run_result in run_results:
        config_data = run_result.get("config", {})  # Still needed for strategy_name
        best_params = run_result.get("best_params", {})
        opt_stats = run_result.get("optimization_stats", {})

        flat_data = {
            "strategy_name": run_result.get(
                "strategy_name"
            ),  # Get strategy name directly from the result dict
            "symbol": run_result.get("symbol"),
            "mode": run_result.get("mode"),
            "target_metric": run_result.get("target_metric"),
        }

        # Extract best params dynamically using all collected keys
        for param_key in sorted_best_param_keys:
            flat_data[param_key] = best_params.get(param_key)  # Use .get() for safety

        # Extract stats
        for stat in REQUESTED_STATS:
            flat_data[stat] = opt_stats.get(stat)

        summary_data_for_excel.append(flat_data)

    results_df = pd.DataFrame(summary_data_for_excel)

    # Define dynamic column order
    column_order = [
        "strategy_name",
        "symbol",
        "mode",
        *sorted_best_param_keys,  # Unpack dynamic best param keys
        "target_metric",
        *REQUESTED_STATS,  # Unpack stats
    ]

    # Ensure all expected columns exist and set order
    for col in column_order:
        if col not in results_df.columns:
            results_df[col] = pd.NA  # Add missing columns with NA
    results_df = results_df[column_order]  # Reorder columns

    return results_df


def _save_dataframe_to_excel(dataframe: pd.DataFrame, excel_filename: str) -> None:
    """
    Saves a Pandas DataFrame to an Excel file.

    Args:
        dataframe: The DataFrame to save.
        excel_filename: The full path for the output Excel file.
    """
    try:
        # Ensure the directory exists before saving
        output_dir = os.path.dirname(excel_filename)
        os.makedirs(output_dir, exist_ok=True)

        # Use ExcelWriter to gain more control
        with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
            dataframe.to_excel(
                writer, index=False, sheet_name="Sheet1"
            )  # Specify sheet name

            # Access the workbook and worksheet
            # workbook = writer.book # Not strictly needed for these operations
            worksheet = writer.sheets["Sheet1"]

            # Freeze the top row (A2 is the first cell below the header)
            worksheet.freeze_panes = "A2"

            # Auto-size columns
            for column_cells in worksheet.columns:
                # Calculate max length, handle None values and header
                try:
                    # Use 0 if cell.value is None, convert others to string
                    # Include header row (index 0) in length calculation
                    length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells
                    )
                except TypeError:
                    # Fallback in case of unexpected types, though str() should handle most
                    length = 10  # Default width if calculation fails

                # Add a little padding
                adjusted_width = length + 2
                # Set column width (using column letter)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Enable filtering for all columns in the header row
            worksheet.auto_filter.ref = worksheet.dimensions
            # --- Highlighting Logic ---
            lime_green_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            highlight_cols = {
                "Equity Final [$]": "max",
                "Commissions [$]": "min",
                "Return [%]": "max",
                "Return (Ann.) [%]": "max",
                "Max. Drawdown [%]": "max",  # Max is closest to zero for negative values
                "Max. Drawdown Duration": "min",
                "Win Rate [%]": "max",
                "Sharpe Ratio": "max",
                "Sortino Ratio": "max",
                "Calmar Ratio": "max",
                "Profit Factor": "max",
            }

            # Get header row (row 1) to map names to column indices
            headers = {cell.value: cell.column for cell in worksheet[1]}

            # --- Per-Symbol Highlighting Logic ---
            best_values_per_symbol = {}
            symbol_col_name = "symbol"
            if symbol_col_name in dataframe.columns:
                grouped_by_symbol = dataframe.groupby(symbol_col_name)

                for col_name, criteria in highlight_cols.items():
                    if col_name in dataframe.columns:
                        best_values_per_symbol[col_name] = {}
                        for symbol, group_df in grouped_by_symbol:
                            numeric_col = pd.to_numeric(
                                group_df[col_name], errors="coerce"
                            ).dropna()
                            if not numeric_col.empty:
                                best_value = (
                                    numeric_col.max()
                                    if criteria == "max"
                                    else numeric_col.min()
                                )
                                best_values_per_symbol[col_name][symbol] = best_value
            # --- End Per-Symbol Highlighting Logic Setup ---

            # Get the column index for 'symbol' in the worksheet
            symbol_col_idx = headers.get(symbol_col_name)

            if symbol_col_idx is not None:
                # Iterate through highlightable columns
                for col_name, criteria in highlight_cols.items():
                    if col_name not in dataframe.columns or col_name not in headers:
                        # Silently skip if column doesn't exist in data or Excel header
                        continue

                    col_idx = headers[col_name]
                    col_letter = get_column_letter(col_idx)

                    # Iterate through worksheet data rows (starting from row 2)
                    for row_idx in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row_idx}"]
                        symbol_cell = worksheet[
                            f"{get_column_letter(symbol_col_idx)}{row_idx}"
                        ]
                        current_symbol = (
                            symbol_cell.value
                        )  # Get the symbol for the current row

                        if current_symbol in best_values_per_symbol.get(col_name, {}):
                            symbol_best_value = best_values_per_symbol[col_name][
                                current_symbol
                            ]

                            try:
                                # Attempt conversion, handle potential errors/None
                                cell_value_numeric = pd.to_numeric(
                                    cell.value, errors="coerce"
                                )

                                # Check if conversion was successful and compare with tolerance
                                if (
                                    pd.notna(cell_value_numeric)
                                    and abs(cell_value_numeric - symbol_best_value)
                                    < 1e-9
                                ):
                                    cell.fill = lime_green_fill
                            except (TypeError, ValueError):
                                # Ignore cells that cannot be converted to numeric
                                pass
            else:
                print(
                    "Warning: 'symbol' column not found in Excel headers. Skipping per-symbol highlighting."
                )

            # --- End Highlighting Logic ---

    except ImportError:
        print(f"Error saving Excel file: Could not import 'openpyxl'.")
        print("Please install it: pip install openpyxl")
    except Exception as e:
        print(f"Error saving Excel file '{excel_filename}': {e}")


def generate_symbol_summary_excel(
    symbol_run_results: List[Dict[str, Any]],
    active_strategy: str,
    symbol: str,
) -> None:
    """
    Processes results for a single symbol and saves them to an Excel file
    in the symbol's optimization_results directory.

    Args:
        symbol_run_results: List of result dictionaries for the specific symbol.
        active_strategy: The name of the strategy being run.
        symbol: The specific symbol these results belong to.
        target_metric: The metric used for optimization for this run.
    """
    if not symbol_run_results:
        print(f"\nNo results for symbol {symbol} to save to Excel.")
        return

    results_df = _process_results_to_dataframe(symbol_run_results)

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Save within the strategy folder
    output_dir = os.path.join("strategies", active_strategy)
    excel_filename = os.path.join(
        output_dir,
        f"{symbol}_{timestamp_str}.xlsx",
    )

    _save_dataframe_to_excel(results_df, excel_filename)


def save_summary_to_excel(
    all_run_results: List[Dict[str, Any]],
    active_strategy: str,
    target_metrics_list: List[str],
) -> None:
    """
    Processes collected optimization results from ALL symbols and saves them
    to a single consolidated Excel file in the main strategy directory.

    Args:
        all_run_results: A list of dictionaries containing results from all symbols.
        active_strategy: The name of the strategy being run.
        target_metrics_list: The list of target metrics used for optimization.
    """
    if not all_run_results:
        print("\nNo consolidated results to save to Excel.")
        return

    # _process_results_to_dataframe already handles extracting the individual target_metric from each run_result
    results_df = _process_results_to_dataframe(all_run_results)

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Save in the main strategy directory
    output_dir = os.path.join("strategies", active_strategy)
    excel_filename = os.path.join(
        output_dir, f"{timestamp_str}_{active_strategy}_summary_.xlsx"
    )

    _save_dataframe_to_excel(results_df, excel_filename)


def generate_overall_summary_excel(
    all_strategies_results: List[Dict[str, Any]],
) -> None:
    """
    Processes collected optimization results from ALL strategies and symbols
    and saves them to a single consolidated Excel file in the 'strategies/' directory.

    Args:
        all_strategies_results: A list of dictionaries containing results
                                 from all strategies and symbols.
    """
    if not all_strategies_results:
        print("\nNo overall results to save to consolidated Excel.")
        return

    # The input list is already flat due to using extend in optimizer_run.py
    # The _process_results_to_dataframe function handles different parameter sets
    # The 'active_strategy' parameter here is less critical as the function
    # primarily uses the 'strategy_name' already within each result dict.
    # The _process_results_to_dataframe function handles different parameter sets
    results_df = _process_results_to_dataframe(all_strategies_results)
    # Sort the DataFrame first by Symbol (asc), then by Return [%] (desc)
    if "symbol" in results_df.columns and "Return [%]" in results_df.columns:
        # Sort by both columns if they exist
        results_df = results_df.sort_values(
            by=["symbol", "Return [%]"], ascending=[True, False]
        )  # Note: False for descending Return [%]
    elif "symbol" in results_df.columns:
        # Fallback: Sort only by Symbol if 'Return [%]' is missing
        results_df = results_df.sort_values(by=["symbol"], ascending=True)
    else:
        # Optional: Log if neither column is present for sorting
        print("WARNING: Could not sort overall summary. 'symbol' column not found.")

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = "strategies"  # Relative path for the strategies directory
    # Ensure the directory exists (handled within _save_dataframe_to_excel)
    excel_filename = os.path.join(
        output_dir, f"{timestamp_str}_overall_optimization_summary.xlsx"
    )

    print(f"Saving overall summary to: {excel_filename}")
    _save_dataframe_to_excel(results_df, excel_filename)
