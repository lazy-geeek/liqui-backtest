"""
Module for writing pandas DataFrames to Excel with formatting and highlighting.
"""

from typing import Dict, Any
import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import numpy as np


def _save_dataframe_to_excel(
    dataframe: pd.DataFrame, excel_filename: str, parameters: Dict[str, Any] = None
) -> None:
    """
    Saves a Pandas DataFrame to an Excel file with formatting:
    - Auto-sized columns
    - Frozen header row
    - Auto-filter
    - Per-symbol best-value and worst-value highlighting
    - Optional parameters sheet

    Args:
        dataframe: The DataFrame to save.
        excel_filename: The full path for the output Excel file.
        parameters: Dictionary of parameters to save in a separate sheet.
    """
    try:
        # Ensure the directory exists before saving
        output_dir = os.path.dirname(excel_filename)
        os.makedirs(output_dir, exist_ok=True)

        with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
            # Write main results sheet
            dataframe.to_excel(writer, index=False, sheet_name="Results")
            worksheet = writer.sheets["Results"]

            # Write parameters sheet if provided
            if parameters:
                params_df = pd.DataFrame(
                    {"Parameter": parameters.keys(), "Value": parameters.values()}
                )
                params_df.to_excel(writer, index=False, sheet_name="Parameters")

                # Format parameters sheet columns
                params_sheet = writer.sheets["Parameters"]
                for column_cells in params_sheet.columns:
                    try:
                        length = max(
                            len(str(cell.value)) if cell.value is not None else 0
                            for cell in column_cells
                        )
                    except TypeError:
                        length = 10
                    adjusted_width = length + 2
                    col_letter = get_column_letter(column_cells[0].column)
                    params_sheet.column_dimensions[col_letter].width = adjusted_width

            # Freeze header row
            worksheet.freeze_panes = "A2"

            # Auto-size columns
            for column_cells in worksheet.columns:
                try:
                    length = max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in column_cells
                    )
                except TypeError:
                    length = 10
                adjusted_width = length + 2
                col_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[col_letter].width = adjusted_width

            # Enable filtering
            worksheet.auto_filter.ref = worksheet.dimensions

            # Prepare highlighting
            lime_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            red_fill = PatternFill(  # New: Light red fill
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )
            highlight_cols = {
                "Equity Final [$]": "max",
                "Commissions [$]": "min",
                "Return [%]": "max",
                "Return (Ann.) [%]": "max",
                "Max. Drawdown [%]": "max",  # Note: For drawdown, max is worst, but we'll handle via criteria
                "Max. Drawdown Duration": "min",
                "Win Rate [%]": "max",
                "Sharpe Ratio": "max",
                "Sortino Ratio": "max",
                "Calmar Ratio": "max",
                "Profit Factor": "max",
            }

            # Map header names to column indices
            headers = {cell.value: cell.column for cell in worksheet[1]}

            # Compute best and worst values per symbol
            best_per_symbol = {}
            worst_per_symbol = {}  # New: For worst values
            symbol_col = "symbol"
            if symbol_col in dataframe.columns:
                grouped = dataframe.groupby(symbol_col)
                for col_name, criteria in highlight_cols.items():
                    if col_name in dataframe.columns:
                        best_per_symbol[col_name] = {}
                        worst_per_symbol[col_name] = {}  # Initialize for worst
                        for symbol, group in grouped:
                            if col_name == "Max. Drawdown Duration":
                                # Handle as Timedelta
                                timedeltas = pd.to_timedelta(
                                    group[col_name], errors="coerce"
                                ).dropna()
                                if not timedeltas.empty:
                                    best = timedeltas.min()  # Smallest duration is best
                                    worst = (
                                        timedeltas.max()
                                    )  # Largest duration is worst
                                    best_per_symbol[col_name][symbol] = best
                                    worst_per_symbol[col_name][symbol] = worst
                            else:
                                numeric = pd.to_numeric(
                                    group[col_name], errors="coerce"
                                ).dropna()
                                if not numeric.empty:
                                    best = (
                                        numeric.max()
                                        if criteria == "max"
                                        else numeric.min()
                                    )
                                    worst = (
                                        numeric.min()
                                        if criteria == "max"
                                        else numeric.max()
                                    )
                                    best_per_symbol[col_name][symbol] = best
                                    worst_per_symbol[col_name][symbol] = worst

            symbol_idx = headers.get(symbol_col)
            if symbol_idx:
                # Apply fills for best and worst values
                for col_name in highlight_cols:
                    if col_name not in headers:
                        continue
                    col_idx = headers[col_name]
                    col_letter = get_column_letter(col_idx)
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col_letter}{row}"]
                        symbol = worksheet[
                            f"{get_column_letter(symbol_idx)}{row}"
                        ].value
                        best_val = best_per_symbol.get(col_name, {}).get(symbol)
                        worst_val = worst_per_symbol.get(col_name, {}).get(symbol)
                        try:
                            if col_name == "Max. Drawdown Duration":
                                # Handle as Timedelta
                                td_val = pd.to_timedelta(cell.value, errors="coerce")
                                if pd.notna(td_val):
                                    if best_val is not None and td_val == best_val:
                                        cell.fill = lime_fill  # Best (min) value
                                    elif worst_val is not None and td_val == worst_val:
                                        cell.fill = red_fill  # Worst (max) value
                            else:
                                val = pd.to_numeric(cell.value, errors="coerce")
                                if pd.notna(val):
                                    if (
                                        best_val is not None
                                        and np.isfinite(val)
                                        and np.isfinite(best_val)
                                        and abs(val - best_val) < 1e-9
                                    ):
                                        cell.fill = lime_fill  # Best value
                                    elif (
                                        worst_val is not None
                                        and np.isfinite(val)
                                        and np.isfinite(worst_val)
                                        and abs(val - worst_val) < 1e-9
                                    ):
                                        cell.fill = red_fill  # Worst value
                        except Exception:
                            continue
            else:
                print(
                    "Warning: 'symbol' column not found. Skipping per-symbol highlighting."
                )
    except ImportError:
        print("Error: failed to import openpyxl. Please install it.")
    except Exception as e:
        print(f"Error saving Excel file '{excel_filename}': {e}")
